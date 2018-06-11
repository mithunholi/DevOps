import openpyxl

wb = openpyxl.load_workbook('./home/mithunkumar/Downloads/Cost_K_PT_SW_MCU_ZBOM-01-R1_0.xls')

# wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet1')
#sheet = wb.get_sheet_by_name('Sheet1')
sheet.cell(row = 10, column = 3).value
c.value
for i in range(10, 105):
print(i, sheet.cell(row=i, column=3).value)
#sheet['A1'].value






wb.save('styled.xlsx')

